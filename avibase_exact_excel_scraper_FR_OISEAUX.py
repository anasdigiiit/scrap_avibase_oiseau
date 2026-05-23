#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Avibase Morocco Exact Excel Scraper
===================================

Scraper requests + BeautifulSoup pour :
1) collecter la checklist Avibase du Maroc ;
2) extraire les fiches espèces, synonymes, noms vernaculaires et distributions ;
3) enrichir les espèces via Oiseaux.net / the-birds.net ;
4) exporter un fichier Excel multi-feuilles ;
5) reprendre automatiquement après interruption avec logs, checkpoints et sauvegardes partielles.

Exécution officielle :
    python avibase_exact_excel_scraper_FR_OISEAUX.py --output output/avibase_maroc_exact.xlsx

Test rapide :
    python avibase_exact_excel_scraper_FR_OISEAUX.py --output output/test.xlsx --limit 10
"""

from __future__ import annotations

import argparse
import logging
import os
from pathlib import Path
import re
import signal
import time
import traceback
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin, urlparse, parse_qs

import requests
from bs4 import BeautifulSoup, Tag
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.checkpoint import CheckpointManager, utc_now_iso
from utils.logger_config import LOGGER_NAME, configure_logger, ensure_directories
from utils.mailer import send_email
from utils.retry_utils import create_retry_session, fetch_text

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

BASE_URL = "https://avibase.bsc-eoc.org"
CHECKLIST_URL = f"{BASE_URL}/checklist.jsp?region=MA&lang=EN"

CRAWL_DELAY = 2.0
REQUEST_TIMEOUT = 120
SAVE_EVERY_DEFAULT = 5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9,fr;q=0.8,ar;q=0.7,es;q=0.6",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

OISEAUX_BASE_URL = "https://www.oiseaux.net"
OISEAUX_LIST_URL = f"{OISEAUX_BASE_URL}/oiseaux/"
THE_BIRDS_BASE_URL = "https://www.the-birds.net"
OISEAUX_DELAY = 0.7

DATA_DIR = Path(
    os.getenv("SCRAPER_DATA_DIR") or os.getenv("RAILWAY_VOLUME_MOUNT_PATH") or str(BASE_DIR)
).expanduser()
LOGS_DIR = DATA_DIR / "logs"
CHECKPOINTS_DIR = DATA_DIR / "checkpoints"
OUTPUT_DIR = DATA_DIR / "output"
LOG_FILE = LOGS_DIR / "scraper.log"
CHECKPOINT_FILE = CHECKPOINTS_DIR / "progress.json"

LOGGER = logging.getLogger(LOGGER_NAME)
SHUTDOWN_REQUESTED = False



@dataclass
class SpeciesRow:
    common_name: str
    scientific_name: str
    checklist_status: str
    avibase_id: str
    species_url: str


# ---------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------

def looks_like_html(html: str) -> bool:
    lowered = html.lower()
    return any(marker in lowered for marker in ("<html", "<body", "<table", "<div", "<p", "<h1", "<h2"))


def fetch_html(
    session: requests.Session,
    url: str,
    delay: float = CRAWL_DELAY,
    context: str = "",
) -> Optional[str]:
    html = fetch_text(
        session,
        url,
        headers=HEADERS,
        delay=delay,
        timeout=REQUEST_TIMEOUT,
        logger=LOGGER,
        context=context or "avibase",
    )
    if html and not looks_like_html(html):
        LOGGER.warning("Response may not be valid HTML for %s | context=%s", url, context or "avibase")
    return html


def handle_shutdown_signal(signum, _frame) -> None:
    global SHUTDOWN_REQUESTED
    SHUTDOWN_REQUESTED = True
    signal_name = signal.Signals(signum).name if signum else f"signal-{signum}"
    LOGGER.warning("Shutdown signal received: %s. The scraper will stop after the current species.", signal_name)


def configure_signal_handlers() -> None:
    signal.signal(signal.SIGTERM, handle_shutdown_signal)
    signal.signal(signal.SIGINT, handle_shutdown_signal)


def soupify(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "lxml")


def clean_text(value: str) -> str:
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_url(href: str, base_url: str = BASE_URL) -> str:
    return urljoin(base_url, href)


# ---------------------------------------------------------------------
# Checklist Maroc
# ---------------------------------------------------------------------

def parse_morocco_checklist(html: str) -> List[SpeciesRow]:
    soup = soupify(html)
    table = soup.find("table", class_="table")
    if table is None:
        return []

    species: List[SpeciesRow] = []

    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) != 3:
            continue

        common_name = clean_text(tds[0].get_text(" "))
        scientific_name = clean_text(tds[1].get_text(" "))
        checklist_status = clean_text(tds[2].get_text(" "))

        link = tds[1].find("a", href=True)
        if not link:
            continue

        href = normalize_url(link["href"])
        qs = parse_qs(urlparse(href).query)
        avibase_id = (qs.get("avibaseid") or [""])[0]

        if not avibase_id:
            continue

        species_url = f"{BASE_URL}/species.jsp?avibaseid={avibase_id}&lang=EN"

        species.append(
            SpeciesRow(
                common_name=common_name,
                scientific_name=scientific_name,
                checklist_status=checklist_status,
                avibase_id=avibase_id,
                species_url=species_url,
            )
        )

    return species


# ---------------------------------------------------------------------
# Species summary page: vernacular names + basics
# ---------------------------------------------------------------------

TARGET_LANGUAGES = {
    "English": "English",
    "French": "French",
    "Arabic": "Arabic",
    "Spanish": "Spanish",
}


def get_page_title(soup: BeautifulSoup) -> str:
    h2 = soup.find("h2")
    if h2:
        return clean_text(h2.get_text(" "))
    return ""


def extract_summary_text(soup: BeautifulSoup) -> str:
    # The first real description paragraph is usually before the image/taxon info.
    for p in soup.find_all("p"):
        txt = clean_text(p.get_text(" "))
        if len(txt) > 120 and "Source:" in txt:
            return txt
    return ""


def extract_taxoninfo_pairs(soup: BeautifulSoup) -> Dict[str, str]:
    """
    Extracts pairs such as:
    Order, Family, Genus, Scientific, Citation, Reference, Protonym, Type locality,
    Avibase ID, Short link, Taxonomic Serial Number.
    """
    result: Dict[str, str] = {}
    taxoninfo = soup.find(id="taxoninfo")
    if not taxoninfo:
        return result

    # The HTML uses <b>Label:</b><br> value
    labels = taxoninfo.find_all("b")
    for b in labels:
        label = clean_text(b.get_text(" ")).rstrip(":")
        if not label:
            continue

        value_parts: List[str] = []
        for sib in b.next_siblings:
            if isinstance(sib, Tag) and sib.name == "b":
                break
            if isinstance(sib, Tag) and sib.name == "hr":
                break
            if isinstance(sib, Tag) and sib.name == "p":
                break
            text = clean_text(sib.get_text(" ") if isinstance(sib, Tag) else str(sib))
            if text:
                value_parts.append(text)

        # Fallback: parent paragraph
        if not value_parts and b.parent:
            parent_text = clean_text(b.parent.get_text(" "))
            parent_text = re.sub(rf"^{re.escape(label)}\s*:?", "", parent_text).strip()
            if parent_text:
                value_parts.append(parent_text)

        value = clean_text(" ".join(value_parts))
        if value:
            key = label.lower().replace(" ", "_")
            if key in result:
                result[key] = f"{result[key]} ; {value}"
            else:
                result[key] = value

    # Short link href
    for a in taxoninfo.find_all("a", href=True):
        href = a["href"]
        if "avibase.ca/" in href:
            result["Lien_court"] = href
        if "itis.gov" in href:
            result["Lien_TSN"] = href

    return result


def extract_vernacular_names(soup: BeautifulSoup, avibase_id: str) -> List[Dict[str, str]]:
    """
    Collects only English, French, Arabic, Spanish.
    Structure observed:
    <b>English:</b><br><span> African Ostrich </span><br>
    """
    rows = []

    for b in soup.find_all("b"):
        label = clean_text(b.get_text(" ")).rstrip(":")
        if label not in TARGET_LANGUAGES:
            continue

        value = ""

        # Usually the next span contains the vernacular name.
        span = b.find_next("span")
        if span:
            value = clean_text(span.get_text(" "))

        if value:
            rows.append(
                {
                    "ID_Avibase": avibase_id,
                    "Langue": TARGET_LANGUAGES[label],
                    "Nom_vernaculaire": value,
                }
            )

    # Deduplicate
    unique = {}
    for r in rows:
        unique[(r["ID_Avibase"], r["Langue"], r["Nom_vernaculaire"])] = r
    return list(unique.values())


# ---------------------------------------------------------------------
# Synonyms page
# ---------------------------------------------------------------------

def extract_original_description_from_synonyms_page(soup: BeautifulSoup) -> Dict[str, str]:
    """
    On the synonyms tab, the top box "Original description" contains:
    Scientific, Citation, Protonym, Reference, Publication, Type locality.
    """
    text = soup.get_text("\n", strip=True)
    keys = ["Scientific", "Citation", "Protonym", "Reference", "Publication", "Type locality"]
    result = {}

    lines = [clean_text(x) for x in text.splitlines() if clean_text(x)]
    for i, line in enumerate(lines):
        label = line.rstrip(":")
        if label in keys and i + 1 < len(lines):
            result[label.lower().replace(" ", "_")] = lines[i + 1]

    # Publication link
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "biodiversitylibrary" in href.lower() or "publication.jsp" in href.lower():
            result.setdefault("Lien_publication", normalize_url(href))
            break

    return result


def parse_synonyms_page(html: str, species: SpeciesRow) -> List[Dict[str, str]]:
    """
    Extracts synonyms isolated in a dedicated sheet.
    Required fields:
    - synonym_name
    - reference
    - citation
    - publication
    - publication_link
    - type_locality

    Avibase pages differ by taxon, so this parser is intentionally tolerant:
    1) It extracts the top "Original description" as a synonym-like source record.
    2) It scans tables for rows that look like scientific synonym names.
    """
    soup = soupify(html)
    rows: List[Dict[str, str]] = []

    original = extract_original_description_from_synonyms_page(soup)

    if original:
        rows.append(
            {
                "ID_Avibase": species.avibase_id,
                "Nom_commun": species.common_name,
                "Nom_scientifique": species.scientific_name,
                "Nom_synonyme": original.get("scientific", ""),
                "Référence": original.get("reference", ""),
                "Citation": original.get("citation", ""),
                "Publication": original.get("publication", ""),
                "Lien_publication": original.get("Lien_publication", ""),
                "Localité_type": original.get("type_locality", ""),
                "Section_source": "Original description",
            }
        )

    # Additional synonym rows from tables.
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            txt = clean_text(tr.get_text(" "))
            if not txt:
                continue

            # Avoid authority list noise and keep rows that appear taxonomic.
            has_italic = tr.find("i") is not None
            if not has_italic:
                continue

            synonym_name = clean_text(tr.find("i").get_text(" ")) if tr.find("i") else ""
            if not synonym_name:
                continue

            # Skip the original scientific name duplicate if already captured,
            # but keep if the table adds additional context.
            rows.append(
                {
                    "ID_Avibase": species.avibase_id,
                    "Nom_commun": species.common_name,
                    "Nom_scientifique": species.scientific_name,
                    "Nom_synonyme": synonym_name,
                    "Référence": "",
                    "Citation": "",
                    "Publication": "",
                    "Lien_publication": "",
                    "Localité_type": "",
                    "Section_source": txt[:500],
                }
            )

    # Deduplicate.
    unique = {}
    for r in rows:
        key = (
            r["ID_Avibase"],
            r.get("Nom_synonyme", ""),
            r.get("Référence", ""),
            r.get("Citation", ""),
            r.get("Publication", ""),
            r.get("Localité_type", ""),
        )
        unique[key] = r
    return list(unique.values())


# ---------------------------------------------------------------------
# MAP / Distribution page
# ---------------------------------------------------------------------

STATUS_ALIASES = {
    "rare/accid.": "rare/accidental",
    "rare/accidental": "rare/accidental",
    "introd.": "introduced",
    "introduced": "introduced",
    "extinct": "extinct/extirpated",
    "extirpated": "extinct/extirpated",
    "endemic": "endemic",
    "present": "present",
}

MOROCCO_ANCHORS = {
    "Morocco",
    "Iriqui National Park",
    "Guelmim - Es-Semara",
    "Tata",
    "Meknès - Tafilalet",
    "Errachidia",
    "Rabat - Salé - Zemmour - Zaer",
    "Skhirate-Témara",
    "Souss - Massa - Draâ",
    "Sous Massa National Park",
    "Chtouka-Ait Baha",
    "Ouarzazate",
    "Inezgane-Ait Melloul",
    "Western Sahara",
}

NON_COUNTRY_GROUPS = {
    "Africa",
    "Continental Africa",
    "Northern Africa",
    "Western Africa",
    "Eastern Africa",
    "Southern Africa",
    "Middle East",
    "Western Palearctic",
    "Palearctic",
    "Holarctic",
    "Senegambia",
    "Mauritania & Senegal",
    "Sudan (incl. South Sudan)",
    "Arabian Peninsula",
}

def normalize_presence(value: str) -> str:
    value = value.strip().lower()
    return STATUS_ALIASES.get(value, value if value else "present")


def parse_name_and_presence(line: str) -> Tuple[str, str]:
    line = clean_text(line)
    m = re.search(r"^(.*?)\s*\[([^\]]+)\]\s*$", line)
    if m:
        return clean_text(m.group(1)), normalize_presence(m.group(2))
    return line, "present"


def extract_distribution_lines(soup: BeautifulSoup) -> List[str]:
    """
    Browser-rendered Avibase map section exposes Distribution as text after the map.
    In HTML, it can be found in the body text. We isolate the zone after 'Distribution'.
    """
    text = soup.get_text("\n", strip=True)
    lines = [clean_text(x) for x in text.splitlines() if clean_text(x)]

    if "Distribution" in lines:
        start = lines.index("Distribution") + 1
    else:
        # fallback: find first line equal ignoring case
        start = 0
        for idx, line in enumerate(lines):
            if line.lower() == "distribution":
                start = idx + 1
                break

    section = lines[start:]

    # Stop before footer/noise.
    stop_markers = [
        "Avibase has been visited",
        "Privacy policy",
        "© Denis Lepage",
    ]
    filtered = []
    for line in section:
        if any(marker in line for marker in stop_markers):
            break
        if line in {"Display regions:", "Countries/territories/regions", "Include states/provinces", "Show all sites"}:
            continue
        if line.startswith("Display regions"):
            continue
        filtered.append(line)

    return filtered


def classify_distribution_entry(name: str, last_country: str) -> str:
    """
    Simple classification:
    - Morocco-related known regions => morocco_region
    - Known high-level biogeographic groups => group
    - Otherwise if previous country exists, many entries can be provinces/sites.
      For Excel, we still collect all entries in country_presence, but mark type.
    """
    if name in MOROCCO_ANCHORS and name != "Morocco":
        return "Région_maroc"
    if name in NON_COUNTRY_GROUPS:
        return "group"
    # Country rows often have no previous region dependency.
    # We keep 'country_or_region' because Avibase mixes countries, territories, provinces and sites.
    return "Pays_ou_région"


def parse_map_distribution(html: str, species: SpeciesRow) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    """
    Returns:
    - countries / broad distribution entries
    - country_presence entries
    - morocco_region entries

    Important: Avibase MAP has colored polygons. Their meanings are visible in the legend:
    present, rare, introduced, endemic, extinct.
    In Distribution text, statuses appear in brackets, e.g. [rare/accid.], [introd.], [extirpated].
    """
    soup = soupify(html)
    lines = extract_distribution_lines(soup)

    countries_rows: List[Dict[str, str]] = []
    presence_rows: List[Dict[str, str]] = []
    morocco_rows: List[Dict[str, str]] = []

    seen_country = set()
    seen_presence = set()
    seen_morocco = set()

    current_country = ""

    for line in lines:
        if not line or line in {"Country", "State/province", "Counties"}:
            continue

        name, presence = parse_name_and_presence(line)
        if not name:
            continue

        entry_type = classify_distribution_entry(name, current_country)

        # update current country when a line is probably country-level
        if entry_type == "Pays_ou_région" and name not in NON_COUNTRY_GROUPS:
            current_country = name

        # Main country/presence sheet: collect entries except high-level groups.
        if entry_type != "group":
            pkey = (species.avibase_id, name, presence)
            if pkey not in seen_presence:
                presence_rows.append(
                    {
                        "ID_Avibase": species.avibase_id,
                        "Nom_commun": species.common_name,
                        "Nom_scientifique": species.scientific_name,
                        "Pays_ou_région": name,
                        "Type_presence": presence,
                        "Source": "species map distribution",
                    }
                )
                seen_presence.add(pkey)

            # countries sheet: keep entries that are likely country/territory rows.
            # It is intentionally broad because Avibase includes countries/territories/regions.
            if entry_type == "Pays_ou_région":
                ckey = (species.avibase_id, name)
                if ckey not in seen_country:
                    countries_rows.append(
                        {
                            "ID_Avibase": species.avibase_id,
                            "Nom_commun": species.common_name,
                            "Nom_scientifique": species.scientific_name,
                            "Pays": name,
                            "Type_presence": presence,
                        }
                    )
                    seen_country.add(ckey)

        if name in MOROCCO_ANCHORS and name != "Morocco":
            mkey = (species.avibase_id, name, presence)
            if mkey not in seen_morocco:
                morocco_rows.append(
                    {
                        "ID_Avibase": species.avibase_id,
                        "Nom_commun": species.common_name,
                        "Nom_scientifique": species.scientific_name,
                        "Région_maroc": name,
                        "Type_presence": presence,
                        "Source": "species map distribution",
                    }
                )
                seen_morocco.add(mkey)

    return countries_rows, presence_rows, morocco_rows



# ---------------------------------------------------------------------
# Oiseaux.net / the-birds.net enrichment
# ---------------------------------------------------------------------

def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value or "")
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def normalize_scientific_name(value: str) -> str:
    return clean_text(value).lower()


def fetch_external_html(session: requests.Session, url: str, delay: float = OISEAUX_DELAY) -> Optional[str]:
    return fetch_html(session, url, delay=delay, context="oiseaux_net")


def extract_family_description_from_oiseaux_page(soup: BeautifulSoup) -> str:
    """Extracts family description from a French Oiseaux.net family page.

    The description is generally placed after the H1 'Famille des ...' and before
    the 'Classement' block. Some families have no description; then returns ''.
    """
    text = soup.get_text("\n", strip=True)
    lines = [clean_text(x) for x in text.splitlines() if clean_text(x)]
    start = None
    for i, line in enumerate(lines):
        if line.startswith("Famille des ") or line.startswith("Famille de "):
            start = i + 1
            break
    if start is None:
        return ""

    desc = []
    skip_patterns = (
        "genre", "genres", "espèce", "espèces", "Classement", "Alphabétique",
        "Systématique", "Filtre", "Fiches", "Photos", "Dessins", "Sons",
        "Nom commun", "Sources :"
    )
    for line in lines[start:]:
        if line.startswith("Classement") or line.startswith("Nom commun"):
            break
        if any(line == p or line.startswith(p) for p in skip_patterns):
            continue
        # A real description line usually contains a sentence.
        if len(line) > 25:
            desc.append(line)
    return clean_text("\n".join(desc))


def discover_oiseaux_family_urls(session: requests.Session) -> List[str]:
    """Discovers family URLs from https://www.oiseaux.net/oiseaux/ without clicks."""
    html = fetch_external_html(session, OISEAUX_LIST_URL, delay=0)
    if not html:
        return []
    soup = soupify(html)
    urls = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = normalize_url(a["href"], OISEAUX_BASE_URL)
        if not href.startswith(f"{OISEAUX_BASE_URL}/oiseaux/"):
            continue
        if not href.endswith(".html"):
            continue
        # Exclude order pages by keeping pages whose link text ends generally with 'idés'.
        label = clean_text(a.get_text(" "))
        if not label:
            continue
        if label.lower().endswith(("idés", "idae", "inés", "dés")) or "idés" in label.lower():
            if href not in seen:
                urls.append(href)
                seen.add(href)
    return urls


def parse_oiseaux_family_page(session: requests.Session, family_url: str) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    """Parses a family page and returns family info + species rows.

    Each species row is matched by exact scientific name and contains the direct
    the-birds.net URL found in the 'Nom anglais' column.
    """
    html = fetch_external_html(session, family_url)
    if not html:
        return {"URL_famille_oiseaux": family_url, "Description_famille": ""}, []
    soup = soupify(html)
    h1 = soup.find("h1")
    family_name_fr = clean_text(h1.get_text(" ")) if h1 else ""
    description = extract_family_description_from_oiseaux_page(soup)
    info = {
        "Nom_famille_oiseaux": family_name_fr,
        "URL_famille_oiseaux": family_url,
        "Description_famille": description,
    }

    species_rows = []
    # Oiseaux.net family table structure: French link, plain scientific name, English link to the-birds.net.
    # We inspect the full table row when available because the scientific name is usually
    # stored in a sibling cell, not in the English-link cell itself.
    for a in soup.find_all("a", href=True):
        href = normalize_url(a["href"], family_url)
        if "the-birds.net/birds/" not in href:
            continue

        scientific = ""
        row = a.find_parent("tr")
        if row:
            cells = row.find_all("td")
            if len(cells) >= 2:
                scientific = clean_text(cells[1].get_text(" "))
            row_text = clean_text(row.get_text(" "))
        else:
            container = a.parent if a.parent else a
            row_text = clean_text(container.get_text(" "))
            if not row_text:
                row_text = clean_text(a.get_text(" "))

        if not scientific:
            m = re.search(r"([A-Z][a-z]+\s+[a-z][a-z\-]+(?:\s+[a-z][a-z\-]+)?)", row_text)
            scientific = clean_text(m.group(1)) if m else ""

        english = clean_text(a.get_text(" "))
        if scientific and english:
            species_rows.append({
                "Nom_scientifique": scientific,
                "Nom_anglais_oiseaux": english,
                "URL_identification_oiseaux": href,
                **info,
            })
    return info, species_rows


def build_oiseaux_scientific_index(session: requests.Session) -> Dict[str, Dict[str, str]]:
    """Builds a scientific-name index from all Oiseaux.net family pages."""
    LOGGER.info("[Oiseaux.net] Discovering family pages...")
    family_urls = discover_oiseaux_family_urls(session)
    LOGGER.info("[Oiseaux.net] Family pages discovered: %s", len(family_urls))
    index: Dict[str, Dict[str, str]] = {}
    for i, family_url in enumerate(family_urls, start=1):
        LOGGER.info("[Oiseaux.net] Family %s/%s: %s", i, len(family_urls), family_url)
        _info, rows = parse_oiseaux_family_page(session, family_url)
        for row in rows:
            key = normalize_scientific_name(row.get("Nom_scientifique", ""))
            if key and key not in index:
                index[key] = row
            elif key and key in index:
                # Keep first but store alternative if needed.
                index[key].setdefault("Autres_URL_possibles", "")
                if row.get("URL_identification_oiseaux") not in index[key]["Autres_URL_possibles"]:
                    index[key]["Autres_URL_possibles"] += (" ; " if index[key]["Autres_URL_possibles"] else "") + row.get("URL_identification_oiseaux", "")
    LOGGER.info("[Oiseaux.net] Species indexed by scientific name: %s", len(index))
    return index


def extract_section_text_by_heading(soup: BeautifulSoup, heading_regex: str, stop_regex: Optional[str] = None) -> str:
    heading = soup.find(lambda tag: tag.name in {"h2", "h3", "h4", "h5"} and re.search(heading_regex, clean_text(tag.get_text(" ")), re.I))
    if not heading:
        return ""
    chunks = []
    for sib in heading.next_siblings:
        if isinstance(sib, Tag) and sib.name in {"h2", "h3", "h4", "h5"}:
            break
        if isinstance(sib, Tag):
            txt = clean_text(sib.get_text(" "))
        else:
            txt = clean_text(str(sib))
        if not txt:
            continue
        if stop_regex and re.search(stop_regex, txt, re.I):
            break
        # Remove play/pause noise but keep call labels.
        if txt.lower() in {"play", "pause"}:
            continue
        chunks.append(txt)
    return clean_text("\n".join(chunks))


def extract_biometrics_as_mensurations(soup: BeautifulSoup) -> str:
    text = extract_section_text_by_heading(soup, r"^Biometrics")
    if not text:
        return ""
    # Normalize common labels into one readable field.
    text = re.sub(r"\bSize\b\s*:?", "Taille :", text, flags=re.I)
    text = re.sub(r"\bWingspan\b\s*:?", "Envergure :", text, flags=re.I)
    text = re.sub(r"\bWeight\b\s*:?", "Poids :", text, flags=re.I)
    return clean_text(text)


def extract_voice_names(soup: BeautifulSoup) -> str:
    section = extract_section_text_by_heading(soup, r"^Voice")
    if not section:
        return ""
    found = []
    for m in re.finditer(r"(?:×♫\s*)?([^\n;]*?\b(?:cri|song|call|chant|vol|alarm|display|duet)[^\n;]*)", section, flags=re.I):
        val = clean_text(m.group(1))
        val = re.sub(r"^(play|pause)\s*", "", val, flags=re.I).strip()
        if val and val not in found:
            found.append(val)
    # Fallback: keep the whole voice section if labels are not detected.
    return "; ".join(found) if found else section


def extract_iucn_status_text(threats_text: str) -> str:
    if not threats_text:
        return ""
    codes = ["EX", "EW", "CR", "EN", "VU", "NT", "LC", "NE"]
    found = [c for c in codes if re.search(rf"\b{c}\b", threats_text)]
    # Usually the active status is graphically highlighted and not easy in text-only mode.
    # We preserve the legend/codes as collected from the section.
    return "; ".join(found) if found else ""


def parse_the_birds_identification_page(session: requests.Session, url: str) -> Dict[str, str]:
    html = fetch_external_html(session, url)
    if not html:
        return {"Erreur_oiseaux": "Page identification Oiseaux.net introuvable"}
    soup = soupify(html)
    threats = extract_section_text_by_heading(soup, r"^Threats")
    return {
        "URL_identification_oiseaux": url,
        "Mensurations": extract_biometrics_as_mensurations(soup),
        "Noms_voix": extract_voice_names(soup),
        "Habitat": extract_section_text_by_heading(soup, r"^Habitat"),
        "Comportement_distinctif": extract_section_text_by_heading(soup, r"^Behaviour"),
        "Vol": extract_section_text_by_heading(soup, r"^Flight"),
        "Notes_sur_Alimentation": extract_section_text_by_heading(soup, r"^Diet"),
        "ReprodNotes": extract_section_text_by_heading(soup, r"^Reproduction"),
        "Répartition_géographique": extract_section_text_by_heading(soup, r"^Geographic range"),
        "Menaces": threats,
        "Statut_conservation_IUCN": extract_iucn_status_text(threats),
        "Sources_bibliographiques": extract_section_text_by_heading(soup, r"^Sources of information"),
        "Erreur_oiseaux": "",
    }


def enrich_species_from_oiseaux_net(session: requests.Session, species: SpeciesRow, oiseaux_index: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    key = normalize_scientific_name(species.scientific_name)
    base = oiseaux_index.get(key)
    if not base:
        return {
            "ID_Avibase": species.avibase_id,
            "Nom_commun": species.common_name,
            "Nom_scientifique": species.scientific_name,
            "Erreur_oiseaux": "Nom scientifique non trouvé dans l'index Oiseaux.net",
        }
    details = parse_the_birds_identification_page(session, base["URL_identification_oiseaux"])
    return {
        "ID_Avibase": species.avibase_id,
        "Nom_commun": species.common_name,
        "Nom_scientifique": species.scientific_name,
        "Nom_anglais_oiseaux": base.get("Nom_anglais_oiseaux", ""),
        "Nom_famille_oiseaux": base.get("Nom_famille_oiseaux", ""),
        "URL_famille_oiseaux": base.get("URL_famille_oiseaux", ""),
        "Description_famille": base.get("Description_famille", ""),
        **details,
    }

# ---------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------

SHEETS = {
    "species": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Statut_checklist_maroc",
        "URL_espece", "Résumé", "Ordre", "Famille", "Genre", "Scientifique",
        "Citation", "Référence", "Protonyme", "Localité_type",
        "Lien_court", "Numéro_taxonomique_sériel", "Lien_TSN"
    ],
    "synonyms": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Nom_synonyme",
        "Référence", "Citation", "Publication", "Lien_publication",
        "Localité_type", "Section_source"
    ],
    "vernacular_names": [
        "ID_Avibase", "Langue", "Nom_vernaculaire"
    ],
    "species_countries": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Pays", "Type_presence"
    ],
    "species_country_presence": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Pays_ou_région", "Type_presence", "Source"
    ],
    "species_morocco_regions": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Région_maroc", "Type_presence", "Source"
    ],
    "oiseaux_net": [
        "ID_Avibase", "Nom_commun", "Nom_scientifique", "Nom_anglais_oiseaux",
        "Nom_famille_oiseaux", "URL_famille_oiseaux", "Description_famille",
        "URL_identification_oiseaux", "Mensurations", "Noms_voix", "Habitat",
        "Comportement_distinctif", "Vol", "Notes_sur_Alimentation",
        "ReprodNotes", "Répartition_géographique", "Menaces",
        "Statut_conservation_IUCN", "Sources_bibliographiques", "Erreur_oiseaux"
    ],
    "errors": [
        "ID_Avibase", "Nom_commun", "URL", "Étape", "Erreur"
    ],
}


def add_rows(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
            width = max(12, min(max_len + 2, 45))
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def export_excel(path: str, all_data: Dict[str, List[Dict[str, str]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        add_rows(ws, headers, all_data.get(sheet_name, []))

    style_workbook(wb)
    wb.save(path)


def build_partial_output_path(output_path: Path) -> Path:
    if output_path.suffix:
        return output_path.with_name(f"{output_path.stem}_partial{output_path.suffix}")
    return output_path.with_name(f"{output_path.name}_partial.xlsx")


def resolve_runtime_path(path_str: str, base_dir: Path = DATA_DIR) -> Path:
    path = Path(path_str).expanduser()
    if path.is_absolute():
        return path
    return base_dir / path


def serialize_species(species: SpeciesRow) -> Dict[str, str]:
    return {
        "ID_Avibase": species.avibase_id,
        "Nom_commun": species.common_name,
        "Nom_scientifique": species.scientific_name,
        "Statut_checklist_maroc": species.checklist_status,
        "URL_espece": species.species_url,
    }


def append_error(
    data: Dict[str, List[Dict[str, str]]],
    species: Optional[SpeciesRow],
    url: str,
    step: str,
    error_message: str,
) -> None:
    row = {
        "ID_Avibase": species.avibase_id if species else "",
        "Nom_commun": species.common_name if species else "",
        "URL": url,
        "Étape": step,
        "Erreur": clean_text(error_message)[:2000],
    }
    data["errors"].append(row)
    LOGGER.warning(
        "Error recorded | step=%s | avibase_id=%s | url=%s | message=%s",
        step,
        row["ID_Avibase"] or "n/a",
        url,
        row["Erreur"],
    )


def save_progress_snapshot(
    checkpoint_manager: CheckpointManager,
    data: Dict[str, List[Dict[str, str]]],
    processed_ids: Set[str],
    stats: Dict[str, Any],
    output_path: Path,
    partial_output_path: Path,
    last_species: Optional[SpeciesRow],
    save_every: int,
    force_excel: bool = False,
) -> None:
    processed_count = len(processed_ids)
    should_export_partial = force_excel or (
        save_every > 0 and processed_count > 0 and processed_count % save_every == 0
    )

    if should_export_partial:
        try:
            export_excel(str(partial_output_path), data)
            stats["partial_exports"] = int(stats.get("partial_exports", 0)) + 1
            LOGGER.info("Partial Excel saved to %s", partial_output_path)
        except Exception as exc:
            LOGGER.error("Partial Excel export failed: %s", exc, exc_info=True)
            append_error(data, last_species, str(partial_output_path), "partial_export", f"Partial export failed: {exc}")

    stats["error_count"] = len(data["errors"])
    stats["updated_at"] = utc_now_iso()

    checkpoint_manager.save(
        data=data,
        processed_ids=processed_ids,
        last_species=serialize_species(last_species) if last_species else None,
        stats=stats,
        output_path=str(output_path),
        partial_output_path=str(partial_output_path),
        completed=bool(stats.get("completed")),
    )


def process_species(
    session: requests.Session,
    species: SpeciesRow,
    data: Dict[str, List[Dict[str, str]]],
    oiseaux_index: Dict[str, Dict[str, str]],
) -> None:
    species_row = serialize_species(species)
    data["species"].append(species_row)

    key_map = {
        "Ordre": "order",
        "Famille": "family",
        "Genre": "genus",
        "Scientifique": "scientific",
        "Citation": "citation",
        "Référence": "reference",
        "Protonyme": "protonym",
        "Localité_type": "type_locality",
        "Lien_court": "Lien_court",
        "Numéro_taxonomique_sériel": "taxonomic_serial_number",
        "Lien_TSN": "Lien_TSN",
    }

    summary_html = fetch_html(session, species.species_url, context=f"summary:{species.avibase_id}")
    if summary_html:
        try:
            soup = soupify(summary_html)
            species_row["Résumé"] = extract_summary_text(soup)

            pairs = extract_taxoninfo_pairs(soup)
            for fr_key, internal_key in key_map.items():
                species_row[fr_key] = pairs.get(internal_key, pairs.get(fr_key, ""))

            data["vernacular_names"].extend(extract_vernacular_names(soup, species.avibase_id))
        except Exception as exc:
            LOGGER.error("Summary parsing failed for %s", species.avibase_id, exc_info=True)
            append_error(data, species, species.species_url, "Résumé", f"Parse error: {exc}")
    else:
        append_error(data, species, species.species_url, "Résumé", "Failed to fetch summary page")

    synonyms_url = f"{BASE_URL}/species.jsp?lang=EN&avibaseid={species.avibase_id}&sec=synonyms"
    synonyms_html = fetch_html(session, synonyms_url, context=f"synonyms:{species.avibase_id}")
    if synonyms_html:
        try:
            data["synonyms"].extend(parse_synonyms_page(synonyms_html, species))
        except Exception as exc:
            LOGGER.error("Synonyms parsing failed for %s", species.avibase_id, exc_info=True)
            append_error(data, species, synonyms_url, "synonyms", f"Parse error: {exc}")
    else:
        append_error(data, species, synonyms_url, "synonyms", "Failed to fetch synonyms page")

    map_url = f"{BASE_URL}/species.jsp?lang=EN&avibaseid={species.avibase_id}&sec=map"
    map_html = fetch_html(session, map_url, context=f"map:{species.avibase_id}")
    if map_html:
        try:
            countries, presence, morocco_regions = parse_map_distribution(map_html, species)
            data["species_countries"].extend(countries)
            data["species_country_presence"].extend(presence)
            data["species_morocco_regions"].extend(morocco_regions)
        except Exception as exc:
            LOGGER.error("Map parsing failed for %s", species.avibase_id, exc_info=True)
            append_error(data, species, map_url, "map", f"Parse error: {exc}")
    else:
        append_error(data, species, map_url, "map", "Failed to fetch map page")

    try:
        oiseaux_row = enrich_species_from_oiseaux_net(session, species, oiseaux_index)
    except Exception as exc:
        LOGGER.error("Oiseaux.net enrichment failed for %s", species.avibase_id, exc_info=True)
        oiseaux_row = {
            "ID_Avibase": species.avibase_id,
            "Nom_commun": species.common_name,
            "Nom_scientifique": species.scientific_name,
            "Erreur_oiseaux": f"Unhandled Oiseaux.net error: {exc}",
        }

    data["oiseaux_net"].append(oiseaux_row)
    if oiseaux_row.get("Erreur_oiseaux"):
        append_error(
            data,
            species,
            oiseaux_row.get("URL_identification_oiseaux", ""),
            "oiseaux_net",
            oiseaux_row.get("Erreur_oiseaux", ""),
        )


def build_notification_body(
    *,
    status: str,
    duration_seconds: float,
    processed_species: int,
    total_species: int,
    output_path: str,
    error_count: int,
    checkpoint_path: str,
    traceback_text: str = "",
) -> str:
    lines = [
        f"Statut: {status}",
        f"Durée (secondes): {duration_seconds:.2f}",
        f"Espèces traitées: {processed_species}/{total_species}",
        f"Fichier de sortie: {output_path}",
        f"Nombre d'erreurs: {error_count}",
        f"Checkpoint: {checkpoint_path}",
    ]
    if traceback_text:
        lines.extend(["", "Traceback:", traceback_text])
    return "\n".join(lines)


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def scrape(
    output: str,
    limit: Optional[int] = None,
    save_every: int = SAVE_EVERY_DEFAULT,
    checkpoint_path: str = str(CHECKPOINT_FILE),
) -> Dict[str, Any]:
    output_path = resolve_runtime_path(output)
    partial_output_path = build_partial_output_path(output_path)
    ensure_directories([LOGS_DIR, CHECKPOINTS_DIR, OUTPUT_DIR, output_path.parent, partial_output_path.parent])

    checkpoint_manager = CheckpointManager(resolve_runtime_path(checkpoint_path), LOGGER)
    checkpoint_payload = checkpoint_manager.load() or {}

    session = create_retry_session()
    started_at = time.monotonic()
    data = {name: [] for name in SHEETS.keys()}
    processed_ids: Set[str] = set()
    last_species: Optional[SpeciesRow] = None

    stats: Dict[str, Any] = {
        "started_at": checkpoint_payload.get("stats", {}).get("started_at", utc_now_iso()),
        "updated_at": utc_now_iso(),
        "partial_exports": int(checkpoint_payload.get("stats", {}).get("partial_exports", 0)),
        "error_count": 0,
        "processed_species": 0,
        "remaining_species": 0,
        "total_species": 0,
        "completed": False,
    }

    if checkpoint_payload and not checkpoint_payload.get("completed"):
        LOGGER.info("Checkpoint detected: %s", checkpoint_manager.path)
        saved_data = checkpoint_payload.get("data", {})
        if isinstance(saved_data, dict):
            for sheet_name in SHEETS:
                rows = saved_data.get(sheet_name, [])
                if isinstance(rows, list):
                    data[sheet_name] = rows
        processed_ids = set(checkpoint_payload.get("processed_avibase_ids", []))
        stats.update({k: v for k, v in checkpoint_payload.get("stats", {}).items() if k in stats})
        LOGGER.info("Resume state loaded | already processed=%s", len(processed_ids))
    elif checkpoint_payload.get("completed"):
        LOGGER.info("Completed checkpoint found at %s. Starting a fresh run.", checkpoint_manager.path)

    LOGGER.info("[1/5] Fetching Morocco checklist...")
    checklist_html = fetch_html(session, CHECKLIST_URL, delay=0, context="morocco_checklist")
    if not checklist_html:
        raise RuntimeError("Cannot fetch Morocco checklist.")

    species_list = parse_morocco_checklist(checklist_html)
    if not species_list:
        raise RuntimeError("Morocco checklist fetched but no species were parsed.")
    if limit:
        species_list = species_list[:limit]

    current_species_ids = {species.avibase_id for species in species_list}
    stale_ids = processed_ids - current_species_ids
    if stale_ids:
        LOGGER.warning("Removing %s stale checkpoint IDs not present in current checklist.", len(stale_ids))
        processed_ids &= current_species_ids

    total_species = len(species_list)
    remaining_species = total_species - len(processed_ids)
    stats["total_species"] = total_species
    stats["processed_species"] = len(processed_ids)
    stats["remaining_species"] = remaining_species
    stats["error_count"] = len(data["errors"])

    LOGGER.info(
        "[2/5] Species found=%s | already processed=%s | remaining=%s",
        total_species,
        len(processed_ids),
        remaining_species,
    )

    save_progress_snapshot(
        checkpoint_manager,
        data,
        processed_ids,
        stats,
        output_path,
        partial_output_path,
        last_species,
        save_every,
        force_excel=False,
    )

    LOGGER.info("[3/5] Building Oiseaux.net index...")
    oiseaux_index = build_oiseaux_scientific_index(session)
    if not oiseaux_index:
        append_error(
            data,
            None,
            OISEAUX_LIST_URL,
            "oiseaux_index",
            "Oiseaux.net scientific index is empty; enrichment may be incomplete.",
        )
    LOGGER.info("[3/5] Oiseaux.net index size: %s", len(oiseaux_index))

    for idx, species in enumerate(species_list, start=1):
        if SHUTDOWN_REQUESTED:
            LOGGER.warning("Graceful shutdown requested before processing next species. Stopping loop.")
            break

        if species.avibase_id in processed_ids:
            continue

        last_species = species
        LOGGER.info(
            "[4/5] %s/%s | %s | %s | avibase_id=%s",
            idx,
            total_species,
            species.common_name,
            species.scientific_name,
            species.avibase_id,
        )

        try:
            process_species(session, species, data, oiseaux_index)
        except Exception as exc:
            LOGGER.error("Unhandled species-level error for %s", species.avibase_id, exc_info=True)
            append_error(data, species, species.species_url, "species", f"Unhandled species error: {exc}")

        processed_ids.add(species.avibase_id)
        stats["processed_species"] = len(processed_ids)
        stats["remaining_species"] = total_species - len(processed_ids)
        stats["error_count"] = len(data["errors"])

        save_progress_snapshot(
            checkpoint_manager,
            data,
            processed_ids,
            stats,
            output_path,
            partial_output_path,
            last_species,
            save_every,
            force_excel=False,
        )

        if SHUTDOWN_REQUESTED:
            LOGGER.warning("Graceful shutdown requested after species %s. Stopping loop.", species.avibase_id)
            break

    if SHUTDOWN_REQUESTED:
        duration_seconds = round(time.monotonic() - started_at, 2)
        stats["completed"] = False
        stats["duration_seconds"] = duration_seconds
        stats["processed_species"] = len(processed_ids)
        stats["remaining_species"] = total_species - len(processed_ids)
        stats["error_count"] = len(data["errors"])

        save_progress_snapshot(
            checkpoint_manager,
            data,
            processed_ids,
            stats,
            output_path,
            partial_output_path,
            last_species,
            save_every,
            force_excel=True,
        )

        LOGGER.warning(
            "Scraping interrupted gracefully | duration=%ss | processed=%s/%s",
            duration_seconds,
            len(processed_ids),
            total_species,
        )
        return {
            "output_path": str(output_path),
            "partial_output_path": str(partial_output_path),
            "checkpoint_path": str(checkpoint_manager.path),
            "stats": stats,
            "error_count": len(data["errors"]),
            "status": "INTERRUPTED",
        }

    LOGGER.info("[5/5] Exporting final Excel to %s", output_path)
    export_excel(str(output_path), data)

    duration_seconds = round(time.monotonic() - started_at, 2)
    stats["completed"] = True
    stats["duration_seconds"] = duration_seconds
    stats["processed_species"] = len(processed_ids)
    stats["remaining_species"] = 0
    stats["error_count"] = len(data["errors"])

    checkpoint_manager.save(
        data=data,
        processed_ids=processed_ids,
        last_species=serialize_species(last_species) if last_species else None,
        stats=stats,
        output_path=str(output_path),
        partial_output_path=str(partial_output_path),
        completed=True,
    )

    LOGGER.info("Scraping completed | duration=%ss | output=%s", duration_seconds, output_path)
    return {
        "output_path": str(output_path),
        "partial_output_path": str(partial_output_path),
        "checkpoint_path": str(checkpoint_manager.path),
        "stats": stats,
        "error_count": len(data["errors"]),
        "status": "SUCCESS",
    }


def main() -> None:
    ensure_directories([LOGS_DIR, CHECKPOINTS_DIR, OUTPUT_DIR])
    configure_logger(LOG_FILE)
    configure_signal_handlers()

    parser = argparse.ArgumentParser(description="Exact Avibase Morocco scraper to Excel.")
    parser.add_argument("--output", required=True, help="Output Excel file, e.g. output/avibase_maroc_exact.xlsx")
    parser.add_argument("--limit", type=int, default=None, help="Optional test limit, e.g. 10")
    parser.add_argument(
        "--save-every",
        type=int,
        default=SAVE_EVERY_DEFAULT,
        help="Save partial Excel every N processed species (default: 5).",
    )
    parser.add_argument(
        "--checkpoint",
        default=str(CHECKPOINT_FILE),
        help="Checkpoint JSON path (default: checkpoints/progress.json).",
    )
    args = parser.parse_args()

    if args.save_every < 1:
        parser.error("--save-every must be >= 1")

    LOGGER.info("Scraper started | output=%s | limit=%s | checkpoint=%s", args.output, args.limit, args.checkpoint)
    LOGGER.info("Runtime data directory: %s", DATA_DIR)
    started_at = time.monotonic()

    try:
        result = scrape(
            output=args.output,
            limit=args.limit,
            save_every=args.save_every,
            checkpoint_path=args.checkpoint,
        )
        status = result.get("status", "SUCCESS")
        if status == "INTERRUPTED":
            LOGGER.warning("Scraper interrupted before completion. Success email will not be sent.")
            return
        body = build_notification_body(
            status=status,
            duration_seconds=float(result["stats"].get("duration_seconds", 0.0)),
            processed_species=int(result["stats"].get("processed_species", 0)),
            total_species=int(result["stats"].get("total_species", 0)),
            output_path=result["output_path"],
            error_count=int(result["error_count"]),
            checkpoint_path=result["checkpoint_path"],
        )
        send_email(
            subject="Avibase Morocco scraper completed",
            body=body,
            attachments=[result["output_path"]],
            logger=LOGGER,
        )
    except Exception as exc:
        failure_duration = round(time.monotonic() - started_at, 2)
        trace_text = traceback.format_exc()
        LOGGER.critical("Critical scraper failure: %s", exc, exc_info=True)

        output_path = resolve_runtime_path(args.output)
        partial_output_path = build_partial_output_path(output_path)
        checkpoint_payload = CheckpointManager(resolve_runtime_path(args.checkpoint), LOGGER).load() or {}
        checkpoint_stats = checkpoint_payload.get("stats", {})
        body = build_notification_body(
            status="FAILED",
            duration_seconds=failure_duration,
            processed_species=int(checkpoint_stats.get("processed_species", 0)),
            total_species=int(checkpoint_stats.get("total_species", 0)),
            output_path=str(output_path),
            error_count=int(checkpoint_stats.get("error_count", 0)),
            checkpoint_path=args.checkpoint,
            traceback_text=trace_text,
        )
        send_email(
            subject="Avibase Morocco scraper failed",
            body=body,
            attachments=[str(partial_output_path), str(LOG_FILE)],
            logger=LOGGER,
        )
        raise


if __name__ == "__main__":
    main()
